import random
import openpyxl
import re
import math

dados = {}

def calculaDistancia(long1,long2,lat1,lat2):
    return math.sqrt(pow(lat2-lat1,2) + pow(long2-long1,2))


def leExcel():

    file = 'dataset.xlsx'
    wb_obj = openpyxl.load_workbook(file)
    sheet_obj = wb_obj.active

    m_row = sheet_obj.max_row

    for i in range(2, m_row + 1):
        latitude = sheet_obj.cell(row=i, column=1).value
        longitude = sheet_obj.cell(row=i, column=2).value
        ponto_recolha = sheet_obj.cell(row=i, column=5).value
        tipo_contentor = sheet_obj.cell(row=i, column=6).value
        contentor_total_litros = int(sheet_obj.cell(row=i, column=10).value)
        contentor_ocupacao = random.randint(1,contentor_total_litros)


        m = re.search(r'(\d+):( )*((\w| )+)',ponto_recolha)
        recolha_id = m.group(1)
        rua = m.group(3)

        m = re.search(r'\((\w+)(\s)*\(',ponto_recolha)
        if m is not None:
            sentido = m.group(1)
        else:
            sentido = ""

        m = re.search(r'\)\:(\s*)(((\w)| )+)(\-?)',ponto_recolha)
        if m is not None:
            ligacao_rua = m.group(2)
        else:
            ligacao_rua = ""

        if (latitude,longitude,recolha_id) not in dados:
            dados[(latitude,longitude,recolha_id)] = (rua,sentido,ligacao_rua,{})
            dados[(latitude, longitude, recolha_id)][3][tipo_contentor] = [contentor_total_litros,contentor_ocupacao]

        elif tipo_contentor not in dados[(latitude, longitude, recolha_id)][3]:
            dados[(latitude, longitude, recolha_id)][3][tipo_contentor] = [contentor_total_litros,contentor_ocupacao]

        elif tipo_contentor in dados[(latitude, longitude, recolha_id)][3]:
            dados[(latitude, longitude, recolha_id)][3][tipo_contentor][0] += contentor_total_litros
            dados[(latitude, longitude, recolha_id)][3][tipo_contentor][1] += contentor_ocupacao

    #print(dados)

def criaGrafo():
    f = open("grafo.pl", "a")

    for key in dados.keys():
        latitude = key[0]
        longitude = key[1]
        ponto = key[2]
        direcao = dados[key][1]

        if direcao=="":
            direcao = "null"
        else:
            direcao = direcao.lower()


        ponto_recolha = "ponto_recolha(" + ponto + ", (" + str(latitude) +", "+ str(longitude) +\
                        "), '" +  dados[key][0] + "', " + direcao + ", ["
        contentores = []

        for contentor in dados[key][3].keys():
            contentores.append("('" + contentor + "', " + str(dados[key][3][contentor][0]) + ", "
                               + str(dados[key][3][contentor][1]) + ")")


        for c in contentores:
            if c == contentores[-1]:
                ponto_recolha = ponto_recolha + c + "]"
            else :
                ponto_recolha = ponto_recolha + c + ", "

        ponto_recolha = ponto_recolha + ")." + "\n"
        f.write(ponto_recolha)


    keys = dict(sorted(dados.items(), key= lambda x: x[0][2]))
    for key in keys:
        rua_conectada = dados[key][2]
        id_conectada = 0

        for keyC in keys:
            if dados[keyC][0] in rua_conectada:
                id_conectada = keyC[2]
                break

        aresta1 = ""
        if id_conectada != 0:
            distancia = calculaDistancia(key[0],keyC[0],key[1],keyC[1])
            aresta1 = "aresta(" + key[2] + ", " + id_conectada + ", " + str(distancia) + ").\n"
            f.write(aresta1)

        keys_list = list(keys)
        ultimo = keys_list[-1]
        if key[2] != ultimo[2]:
            next_key = keys_list[keys_list.index(key)+1]
            distancia = calculaDistancia(key[0],next_key[0],key[1],next_key[1])
            aresta2 = "aresta(" + key[2] + ", " + next_key[2] + ", " + str(distancia) + ").\n"

            if (aresta2 != aresta1):
                f.write(aresta2)


    f.close()


if __name__ == '__main__':
    dist1 =calculaDistancia(-9.14301166531656, -9.14329884772974, 38.7073877536231, 38.7065565483955)
    dist2 = calculaDistancia(-9.14301166531656, -9.14473414894124, 38.7073877536231, 38.7064360246404)
    print(str(dist1) + "   " + str(dist2))
    leExcel()
    criaGrafo()